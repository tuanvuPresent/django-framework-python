import csv
from textwrap import wrap
from zipfile import BadZipFile

import openpyxl as openpyxl
import xlwt as xlwt
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from drf_yasg.utils import swagger_auto_schema
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet

from apps.account.models import User
from apps.common.constant import ErrorCode
from apps.common.custom_exception_handler import CustomAPIException
from apps.exportfile.export_service import ExportService
from apps.common.custom_model_view_set import BaseGenericViewSet
from apps.common.serializer import NoneSerializer
from apps.exportfile.serializer import CVSFileSerializer
from apps.e_learning.models.question import Question
from apps.upfile.serializer import FileSerializer


class ExportAPIView(GenericViewSet):
    serializer_class = NoneSerializer

    @action(methods=['GET'], detail=False)
    def export_users_xls(self, request):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="users.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Users')

        # Sheet header, first row
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        columns = ['pk', 'username', 'first_name', 'last_name', 'gender', 'password']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()
        rows = User.objects.all().values_list('pk', 'username', 'first_name', 'last_name', 'gender', 'password')
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
        wb.save(response)
        return response

    @action(methods=['get'], detail=False)
    def export_users_csv(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="users.csv"'

        writer = csv.writer(response)
        writer.writerow(['Username', 'First name', 'Last name', 'Email address'])

        users = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
        for user in users:
            writer.writerow(user)

        return response

    # @action(methods=['get'], detail=False, url_path='effort-xlsx')
    # def export_xlsx(self, request):
    #     data = {
    #         'Column 1': [],
    #         'Column 2': [],
    #     }
    #     file_name = 'effort.xlsx'
    #     df = pd.DataFrame(data)
    #     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #     response['Content-Disposition'] = 'attachment; filename=""' + file_name
    #     df.to_excel(response, index=False)
    #     return response


class ExportV2APIView(GenericViewSet):

    @action(methods=['GET'], detail=False)
    def export_users_xls(self, request):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="users.xls"'
        
        response = ExportService().export(response)
        return response

class ImportAPIView(BaseGenericViewSet):
    parser_classes = (MultiPartParser,)

    @action(methods=['POST'], detail=False)
    @swagger_auto_schema(request_body=FileSerializer)
    def import_xlsx(self, request):
        excel_file = request.FILES["file"]
        try:
            wb = openpyxl.load_workbook(excel_file)
        except BadZipFile:
            raise CustomAPIException(**ErrorCode.FORMAT_FILE)

        excel_data = list()
        for worksheet in wb:
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

        return Response(data=excel_data)

    @action(methods=['POST'], detail=False)
    @swagger_auto_schema(request_body=CVSFileSerializer)
    def import_csv(self, request):
        file = request.FILES
        serializer = CVSFileSerializer(data=file)
        serializer.is_valid(raise_exception=True)

        file = file['file']
        lines = file.read().decode('UTF-8').split('\n')
        lines.pop(len(lines) - 1)

        items = ['type', 'name', 'pass', 'No']
        data = []
        for line in lines:
            res = {}
            line = line.split(',')
            col = 0
            for item in items:
                res[str(item)] = line[col]
                col += 1
            data.append(res)
        return Response(data=data)

    # @action(methods=['post'], detail=False, url_path='import-csv-xlsx', parser_classes=[MultiPartParser, ])
    # @swagger_auto_schema(request_body=CVSFileSerializer)
    # @transaction.atomic()
    # def import_csv_or_xlsx(self, request):
    #     file = request.FILES['file']
    #     try:
    #         if file.content_type == 'text/csv':
    #             df = pd.read_csv(file)
    #         else:
    #             df = pd.read_excel(file)
    #     except:
    #         raise CustomAPIException()
    #     row_iter = df.iterrows()
    #     for index, row in row_iter:
    #         pass
    #     return Response()

class ExportPdfAPIView(GenericViewSet):
    serializer_class = NoneSerializer

    @action(methods=['GET'], detail=False)
    def write_pdf_1(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="somefilename.pdf"'

        p = canvas.Canvas(response)
        p.drawString(100, 100, "Hello world.")
        p.showPage()
        p.save()

        return response

    @action(methods=['GET'], detail=False)
    def write_pdf_2(self, request):
        doc = SimpleDocTemplate("/tmp/download.pdf", pagesizes=A4)
        styles = getSampleStyleSheet()
        Story = []
        style = styles["Normal"]
        for i in range(10):
            bogustext = ("This is This number This is This numberThis is This numberThis is This numberThis is "
                         "This numberThis is This numberThis is This numberThis is This numberThis is This "
                         "numberThis is This numberThis is This numberThis is This number %s.  ---- " % i) * 20
            p = Paragraph(bogustext, style)
            Story.append(p)
            Story.append(Spacer(1, 0.2 * inch))
        doc.build(Story)

        fs = FileSystemStorage("/tmp")
        with fs.open("download.pdf") as pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="download.pdf"'
            return response

    @action(methods=['GET'], detail=False)
    def write_pdf_3(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="download.pdf"'

        pdf = canvas.Canvas(response)
        pdf.setTitle('Title')
        pdf.drawString(200, 800, 'DANH SACH CAU HOI')
        question = Question.objects.prefetch_related('answers')
        x = 40
        y = 750
        for i in range(10):
            for item in question:
                pdf.drawString(x, y, item.content + ':')
                y -= 15
                if y < 40:
                    pdf.showPage()
                    y = 800
                for index, answer in enumerate(item.answers.all()):
                    pdf.drawString(x, y, '    ' + str(index) + '.' + answer.content)
                    y -= 15
                    if y < 40:
                        pdf.showPage()
                        y = 800

        text = "Lorem Ipsum is simply dummy text of the printing and typesetting industry. " \
               "Lorem Ipsum has been the industry's standard dummy text ever since the 1500s, " \
               "when an unknown printer took a galley of type and scrambled it to make a type specimen book."
        t = pdf.beginText(40, 800)
        wraped_text = "\n".join(wrap(text, 80))  # 80 is line width
        pdf.showPage()
        t.textLines(wraped_text)
        pdf.drawText(t)

        pdf.save()
        return response
